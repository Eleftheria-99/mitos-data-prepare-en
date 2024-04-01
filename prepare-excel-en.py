import openpyxl

# Load your existing Excel file
workbook = openpyxl.load_workbook('MITOS-EN-QAS.xlsx')
sheet = workbook.active

# Create a new workbook to store modified data
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active

# Loop through each cell in your Greek data
# print(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)[0])
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    if row[0].value != None: 
        original_row = row[0].value
        print("row : " + original_row + "\n")

        modified_row = "<s>[INST]" + original_row + "</s>"
        splitted_row = modified_row.split("\n")

        if len(splitted_row) >= 2: 
            modified_row = splitted_row[0] + "[/INST]" + splitted_row[1]
            print("final row : " + modified_row + "\n")
        
        # Write the modified row to the new sheet
        new_sheet.append([modified_row])

        # Save the new workbook
        new_workbook.save('modified_rows.xlsx')

print(f"done\n")
